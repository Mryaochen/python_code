#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/5/3 10:02
# @Author  : name
# @Site    : 
# @File    : 练习题：selenium登录QQ邮箱获取邮件title.py
# @Software: PyCharm
from selenium import webdriver
from lxml import etree
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from openpyxl import Workbook
if __name__ == '__main__':
    path = r'./drivers/geckodriver.exe'
    url = 'https://mail.qq.com/?cancel_login=true&from=session_timeout'
    binary = FirefoxBinary(r'D:\installation\Mozilla Firefox\firefox.exe')
    driver = webdriver.Firefox(executable_path=path,firefox_binary=binary)
    driver.get(url)
    time.sleep(1)
    driver.switch_to.frame(driver.find_element(By.CSS_SELECTOR,'.QQMailSdkTool_login_loginBox_qq_iframe'))
    driver.switch_to.frame('ptlogin_iframe')  # 切换浏览器标签定位的作用域
    button = driver.find_element(By.CSS_SELECTOR,'#switcher_plogin')
    button.click()
    time.sleep(5)

    usr_input = driver.find_element(By.ID,'u')
    pwd_input = driver.find_element(By.ID,'p')
    login_button = driver.find_element(By.ID,'login_button')
    QQ = input('QQ:')
    PWD = input('QQ密码：')
    usr_input.send_keys(QQ)
    time.sleep(1)
    pwd_input.send_keys(PWD)
    time.sleep(1)
    login_button.click()
    time.sleep(5)
    driver.switch_to.default_content()
    # driver.save_screenshot('qq.png')
    INBOX = driver.find_element(By.ID,'folder_1')
    INBOX.click()
    time.sleep(1)
    driver.switch_to.frame('mainFrame') #切换ifarme 否则页面原数据获取不到所需元素
    page_text = driver.page_source
    tree = etree.HTML(page_text)
    div_list = tree.xpath('//div[@id = "div_showbefore"]/table//div[@class="tf no"]')
    print(div_list)
    title_list = []
    for div in  div_list:
        title = div.xpath('./u/text()')[0]
        title_list.append(title)
    print(title_list)
    time.sleep(5)
    driver.quit()
    #写入excel
    TargetPath = 'title.xlsx'
    wb = Workbook()
    sheet = wb.create_sheet('post_title',0)
    title_name = sheet.cell(row=1,column=1)
    title_name.value = '邮件名'
    for i in range(len(title_list)):
        title_name = sheet.cell(column = 1,row = i+2)
        title_name.value = title_list[i]
    wb.save(TargetPath)

